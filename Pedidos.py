import pyperclip
import pyautogui as pg
import openpyxl as xl
import time

wb = xl.load_workbook('C:\\Users\\gabur\\OneDrive\\Desktop\\Concat.xlsx')
ws = wb['Saldo Estoque']


# Abre o SLD

pg.alert('Cuidado, Não mexer no computador')
pg.PAUSE = 0.5
pg.press('winleft')
pg.write('edge')
pg.press('enter')
pg.moveTo(1848, 52)
pg.click()
pg.moveTo(1728, 279)
pg.click()
pg.moveTo(540, 78)
pg.click()
time.sleep(3)
pg.moveTo(951, 494)
pg.click()
time.sleep(75)
pg.moveTo(1275, 239)
pg.click()

pg.PAUSE = 0.3
pg.moveTo(15, 108)
pg.click()

for row in range(2, 4):
    if ws['F' + str(row)].value == "B":
        pg.moveTo(624, 368)
        pg.click()
        pg.hotkey('ctrl', 'a')
        pg.press('backspace')
        pg.write(ws['A' + str(row)].value)
        pg.moveTo(769, 365)
        pg.click()  #pesquisa
        pg.moveTo(1086, 430)
        pg.mouseDown()
        pg.moveTo(910, 430)
        pg.mouseUp()
        pg.hotkey('ctrl', 'c')
        ws['G' + str(row)].value = pyperclip.paste()
        print(ws['G' + str(row)].value)
    if ws['A' + str(row)].value == None:
        break

pg.hotkey('alt', 'tab')


for row in range(6, 100):
    if ws['F' + str(row)].value == "B":
        print(ws['G' + str(row)].value)
    if ws['A' + str(row)].value == None:
        break


wb.save('C:\\Users\\gabur\\OneDrive\\Desktop\\Concat.xlsx')
